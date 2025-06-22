# -----------------------------------------------------------------------------
# ResearchAgent Tool: Multi-Format File Querying (Phase 15 - FIX 2)
#
# FIX: This version improves file handling based on user feedback.
# 1.  The `else` block now attempts to read any unrecognized file extension as
#     plain text instead of skipping it.
# 2.  This makes the tool more versatile for formats like .csv, .md, .py, etc.
# -----------------------------------------------------------------------------

import os
import logging
from typing import List

# --- Document Parsing Libraries ---
import pypdf
import docx
import openpyxl

# --- LangChain & Pydantic Core ---
from langchain_core.tools import StructuredTool
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_community.chat_models import ChatOllama

# --- Local Imports ---
from .file_system import _resolve_path

logger = logging.getLogger(__name__)


# --- Helper Functions for Text Extraction ---

def _read_txt(path: str) -> str:
    """Reads text from a plain .txt file."""
    with open(path, 'r', encoding='utf-8', errors='ignore') as f:
        return f.read()

def _read_pdf(path: str) -> str:
    """Extracts text from all pages of a .pdf file."""
    try:
        reader = pypdf.PdfReader(path)
        content = [page.extract_text() for page in reader.pages]
        return "\n".join(content)
    except Exception as e:
        logger.error(f"Error reading PDF {os.path.basename(path)}: {e}")
        return f"[Error reading PDF: {e}]"

def _read_docx(path: str) -> str:
    """Extracts text from all paragraphs of a .docx file."""
    try:
        document = docx.Document(path)
        content = [p.text for p in document.paragraphs]
        return "\n".join(content)
    except Exception as e:
        logger.error(f"Error reading DOCX {os.path.basename(path)}: {e}")
        return f"[Error reading DOCX: {e}]"

def _read_xlsx(path: str) -> str:
    """Extracts text from all cells in all sheets of an .xlsx file."""
    try:
        workbook = openpyxl.load_workbook(path, data_only=True)
        full_content = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_content = []
            for row in sheet.iter_rows():
                row_content = [str(cell.value) if cell.value is not None else "" for cell in row]
                sheet_content.append("\t".join(row_content))
            full_content.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(sheet_content))
        return "\n\n".join(full_content)
    except Exception as e:
        logger.error(f"Error reading XLSX {os.path.basename(path)}: {e}")
        return f"[Error reading XLSX: {e}]"


# --- Core Tool Logic ---

def query_files(files: List[str], question: str, workspace_path: str) -> str:
    """
    Reads content from a list of files (PDF, DOCX, XLSX, TXT) and uses an LLM
    to answer a specific question based on their combined content.
    """
    aggregated_content = []
    
    if not files:
        return "Error: No files were provided to query."

    logger.info(f"Querying {len(files)} files with question: '{question}'")

    for file_path in files:
        try:
            full_path = _resolve_path(workspace_path, file_path)
            filename = os.path.basename(full_path)
            
            content = ""
            if filename.lower().endswith(".pdf"):
                content = _read_pdf(full_path)
            elif filename.lower().endswith(".docx"):
                content = _read_docx(full_path)
            elif filename.lower().endswith(".xlsx"):
                content = _read_xlsx(full_path)
            else:
                # --- MODIFIED: Attempt to read any other file as text ---
                logger.info(f"Unsupported extension for '{filename}', attempting to read as plain text.")
                content = _read_txt(full_path)


            aggregated_content.append(f"--- Content from: {filename} ---\n{content}")

        except FileNotFoundError:
            aggregated_content.append(f"[Error: File not found at '{file_path}']")
        except Exception as e:
            logger.error(f"Failed to read or process file '{file_path}': {e}", exc_info=True)
            aggregated_content.append(f"[Error processing file '{file_path}': {e}]")
    
    combined_context = "\n\n".join(aggregated_content)

    if not combined_context.strip():
        return "Error: Could not extract any content from the provided files."
    
    # --- LLM Synthesis Step ---
    try:
        llm_id = os.getenv("EDITOR_LLM_ID", "gemini::gemini-1.5-pro-latest")
        provider, model_name = llm_id.split("::")
        logger.info(f"query_files tool is using Editor LLM: {llm_id}")

        if provider == "gemini":
            llm = ChatGoogleGenerativeAI(model=model_name, google_api_key=os.getenv("GOOGLE_API_KEY"))
        elif provider == "ollama":
            llm = ChatOllama(model=model_name, base_url=os.getenv("OLLAMA_BASE_URL"))
        else:
            return f"Error: Unsupported LLM provider '{provider}' specified in EDITOR_LLM_ID."

        prompt = f"""
You are an expert research assistant. Your task is to answer a specific question based ONLY on the provided text context from one or more documents.

**User's Question:**
{question}

**Provided Context from Files:**
---
{combined_context}
---

Based solely on the context above, provide a comprehensive answer to the user's question. If the context does not contain the answer, state that clearly. Do not use any external knowledge.
"""
        response = llm.invoke(prompt)
        logger.info("Successfully synthesized an answer using the Editor LLM.")
        return response.content

    except Exception as e:
        logger.error(f"An error occurred during LLM synthesis: {e}", exc_info=True)
        return f"Error: Failed to synthesize an answer. The LLM returned an error: {e}"


# --- Tool Definition ---

tool = StructuredTool.from_function(
    func=query_files,
    name="query_files",
    description="A powerful tool that reads content from multiple files of different formats (.pdf, .docx, .xlsx, .txt) and answers a specific question based on their combined content. Use this for summarization, comparison, data extraction, and analysis across documents."
)